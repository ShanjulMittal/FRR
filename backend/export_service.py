import pandas as pd
import io
import json
import csv
import ipaddress
from datetime import datetime
from typing import List, Dict, Any, Optional, Iterator
from models import db, ReviewResult, NormalizedRule, ComplianceRule, ReviewProfile, CMDBAsset
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

def generate_excel_export(review_session_id: str, include_compliant: Optional[bool] = None) -> bytes:
    """
    Generate Excel export with compliance status columns for a review session.
    Returns the Excel file as bytes.
    """
    try:
        from openpyxl import Workbook

        include = True if include_compliant is None else bool(include_compliant)

        base_query = db.session.query(ReviewResult).options(
            joinedload(ReviewResult.normalized_rule).joinedload(NormalizedRule.raw_rule),
            joinedload(ReviewResult.compliance_rule),
            joinedload(ReviewResult.profile)
        ).filter(ReviewResult.review_session_id == review_session_id)
        if not include:
            base_query = base_query.filter(ReviewResult.status == 'non_compliant')
        base_query = base_query.join(NormalizedRule).order_by(NormalizedRule.source_file.asc(), ReviewResult.id.asc())

        has_any = db.session.query(base_query.exists()).scalar()
        if not has_any:
            raise ValueError("No results found for the specified review session")

        assets = db.session.query(CMDBAsset.ip_address, CMDBAsset.hostname, CMDBAsset.additional_data).all()
        ip_pci_map: Dict[str, str] = {}
        host_pci_map: Dict[str, str] = {}
        for ip, hostname, data in assets:
            if not data:
                continue
            try:
                d = json.loads(data)
                cat = d.get('pcidss_asset_category')
                if not cat:
                    continue
                if ip:
                    ip_pci_map[str(ip).strip().lower()] = cat
                if hostname:
                    host_pci_map[str(hostname).strip().lower()] = cat
            except Exception:
                continue

        def get_pci_categories(ip_field: str, hostname_field: Optional[str] = None) -> str:
            cats = set()
            if ip_field:
                tokens = [t.strip() for t in str(ip_field).replace(';', ',').split(',') if t.strip()]
                for t in tokens:
                    t_clean = t.lower().replace('host ', '').strip()
                    cat = ip_pci_map.get(t_clean)
                    if cat:
                        cats.add(cat)
            if hostname_field:
                h_clean = str(hostname_field).lower().strip()
                cat = host_pci_map.get(h_clean)
                if cat:
                    cats.add(cat)
            return ', '.join(sorted(cats))

        headers = [
            'Rule_ID','Rule_Name','Line_Number','Action','Protocol','Source_IP','Source_Zone','Source_Port',
            'Dest_IP','Dest_Zone','Dest_Port','Service_Name','Source_VLAN','Dest_VLAN','Interface','Direction',
            'Logging','Description','Original_Rule','Raw_Text','Raw_Rule_Text','Source_PCI_DSS_Categories',
            'Dest_PCI_DSS_Categories','Compliance_Status','Failed_Checks','Severity','Compliance_Rule_Name',
            'Check_Description','Field_Checked','Operator','Expected_Value','Actual_Value','Non_Compliance_Type',
            'Findings_Details','Notes','Checked_At','Plain_Explanation','Why_It_Matters','Raw_Rule_Details'
        ]

        wb = Workbook(write_only=True)

        summary_ws = wb.create_sheet(title='Summary')
        dashboard_ws = wb.create_sheet(title='Dashboard')
        sev_ws = wb.create_sheet(title='Severity Breakdown')
        top_ws = wb.create_sheet(title='Top Violations')

        file_ws_by_key: Dict[str, Any] = {}
        file_title_by_key: Dict[str, str] = {}
        used_titles: Dict[str, int] = {}

        def safe_sheet_title(name: str) -> str:
            base = (name or 'Unknown').replace('/', '_').replace('\\', '_')
            base = ''.join(ch for ch in base if ch not in '[]:*?/\\')
            base = base.strip() or 'Unknown'
            base = base[:31]
            if base not in used_titles:
                used_titles[base] = 1
                return base
            used_titles[base] += 1
            suffix = f"_{used_titles[base]}"
            return (base[:31 - len(suffix)] + suffix)[:31]

        file_stats: Dict[str, Dict[str, Any]] = {}
        total_rules = 0
        total_compliant = 0
        total_non_compliant = 0
        severity_breakdown: Dict[str, int] = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
        top_violations: Dict[str, int] = {}

        def ensure_ws_for_source(source_file: str):
            if source_file in file_ws_by_key:
                return file_ws_by_key[source_file]
            title = safe_sheet_title(source_file)
            ws = wb.create_sheet(title=title)
            ws.append(headers)
            file_ws_by_key[source_file] = ws
            file_title_by_key[source_file] = title
            return ws

        for result in base_query.yield_per(1000):
            nr = result.normalized_rule
            source_file = getattr(nr, 'source_file', '') or 'Unknown'
            ws = ensure_ws_for_source(source_file)

            file_stat = file_stats.setdefault(source_file, {'total': 0, 'compliant': 0, 'non_compliant': 0})
            file_stat['total'] += 1
            total_rules += 1

            status_value = (result.status or '').replace('_', ' ').title()
            if status_value == 'Compliant':
                file_stat['compliant'] += 1
                total_compliant += 1
            else:
                file_stat['non_compliant'] += 1
                total_non_compliant += 1

            sev = result.severity or 'Medium'
            if sev in severity_breakdown:
                severity_breakdown[sev] += 1

            vname = result.compliance_rule.rule_name if result.compliance_rule else ''
            if vname and status_value != 'Compliant':
                top_violations[vname] = top_violations.get(vname, 0) + 1

            safe = lambda name: getattr(nr, name, '')
            src_cats = get_pci_categories(getattr(nr, 'source_ip', ''), getattr(nr, 'source_hostname', None))
            dst_cats = get_pci_categories(getattr(nr, 'dest_ip', ''), getattr(nr, 'dest_hostname', None))

            checks_list: List[Dict[str, Any]] = []
            try:
                if isinstance(result.failed_checks, str):
                    parsed = json.loads(result.failed_checks) if result.failed_checks else []
                else:
                    parsed = result.failed_checks or []
                checks_list = parsed if isinstance(parsed, list) else []
            except Exception:
                checks_list = []

            field_checked = ''
            op = ''
            expected = ''
            actual = ''
            plain_expl = ''
            why = ''
            findings_details = ''
            if checks_list:
                first = checks_list[0] if isinstance(checks_list[0], dict) else {}
                op = (first.get('operator') or '').lower()
                expected = first.get('expected_value') or ''
                actual = first.get('actual_value') or ''
                field_checked = first.get('field_checked') or ''
                meaning = {
                    'equals': f"should equal '{expected}' but is '{actual}'",
                    'not_equals': f"should not equal '{expected}' but is '{actual}'",
                    'contains': f"should include '{expected}' but is '{actual}'",
                    'not_contains': f"should not include '{expected}' but is '{actual}'",
                    'in_list': f"should be one of '{expected}' but is '{actual}'",
                    'not_in_list': f"should not be any of '{expected}' but is '{actual}'",
                    'regex_match': f"should match pattern '{expected}' but is '{actual}'",
                    'not_regex_match': f"should not match pattern '{expected}' but is '{actual}'",
                    'starts_with': f"should start with '{expected}' but is '{actual}'",
                    'ends_with': f"should end with '{expected}' but is '{actual}'",
                    'is_empty': f"should be empty but is '{actual}'",
                    'is_not_empty': "should not be empty",
                    'greater_than': f"should be > '{expected}' but is '{actual}'",
                    'greater_than_or_equal': f"should be >= '{expected}' but is '{actual}'",
                    'less_than': f"should be < '{expected}' but is '{actual}'",
                    'less_than_or_equal': f"should be <= '{expected}' but is '{actual}'",
                    'composite': "did not satisfy combined conditions"
                }
                plain_expl = meaning.get(op, f"expected '{expected}' using '{op}', actual '{actual}'")
                why = f"Severity {result.severity or ''}: higher severity needs faster fix."
                try:
                    findings_details = json.dumps(checks_list)
                except Exception:
                    findings_details = ''

            raw_rule_text = ''
            raw_rule_details = ''
            try:
                raw_obj = getattr(nr, 'raw_rule', None)
                if raw_obj:
                    raw_rule_text = raw_obj.rule_text or raw_obj.raw_text or ''
                    raw_dict = {
                        'id': raw_obj.id,
                        'rule_name': raw_obj.rule_name,
                        'source_file': raw_obj.source_file,
                        'line_number': raw_obj.file_line_number,
                        'action': raw_obj.action,
                        'protocol': raw_obj.protocol,
                        'source': raw_obj.source,
                        'destination': raw_obj.destination,
                        'service': raw_obj.destination,
                        'raw_text': raw_obj.raw_text
                    }
                else:
                    raw_dict = {'id': safe('id'), 'rule_name': safe('rule_name'), 'raw_text': safe('raw_text')}
                raw_rule_details = json.dumps(raw_dict)
            except Exception:
                raw_rule_text = ''
                raw_rule_details = ''

            row = [
                safe('id'),
                safe('rule_name'),
                safe('line_number'),
                safe('action'),
                safe('protocol'),
                safe('source_ip'),
                safe('source_zone'),
                safe('source_port'),
                safe('dest_ip'),
                safe('dest_zone'),
                safe('dest_port'),
                safe('service_name'),
                safe('source_vlan'),
                safe('dest_vlan'),
                safe('interface'),
                safe('direction'),
                safe('logging'),
                safe('description'),
                safe('original_rule'),
                safe('raw_text'),
                raw_rule_text,
                src_cats,
                dst_cats,
                status_value,
                len(checks_list),
                result.severity,
                vname,
                result.compliance_rule.description if result.compliance_rule else '',
                field_checked,
                op,
                expected,
                actual,
                vname,
                findings_details,
                result.notes or '',
                result.checked_at.strftime('%Y-%m-%d %H:%M:%S') if result.checked_at else '',
                plain_expl,
                why,
                raw_rule_details
            ]
            ws.append(row)

        summary_ws.append(['Source_File', 'Total_Rules', 'Compliant', 'Non_Compliant', 'Compliance_Percentage'])
        for source_file, st in sorted(file_stats.items(), key=lambda x: x[0]):
            total = int(st.get('total') or 0)
            comp = int(st.get('compliant') or 0)
            nonc = int(st.get('non_compliant') or 0)
            pct = round((comp / total) * 100, 2) if total else 0
            summary_ws.append([source_file, total, comp, nonc, pct])
        overall_pct = round((total_compliant / total_rules) * 100, 2) if total_rules else 0
        summary_ws.append(['OVERALL TOTAL', total_rules, total_compliant, total_non_compliant, overall_pct])

        dashboard_ws.append(['Metric', 'Value'])
        dashboard_ws.append(['Total Rules', total_rules])
        dashboard_ws.append(['Compliant Rules', total_compliant])
        dashboard_ws.append(['Non-Compliant Rules', total_non_compliant])
        dashboard_ws.append(['Compliance %', overall_pct])

        sev_ws.append(['Severity', 'Count'])
        for k in ['Critical', 'High', 'Medium', 'Low']:
            sev_ws.append([k, severity_breakdown.get(k, 0)])

        top_ws.append(['Rule Name', 'Violations'])
        for name, count in sorted(top_violations.items(), key=lambda x: x[1], reverse=True)[:10]:
            top_ws.append([name, count])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        raise Exception(f"Error generating Excel export: {str(e)}")

def generate_excel_export_custom(review_session_id: str, options: Dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook

        source_file = options.get('source_file')
        include_compliant_opt = options.get('include_compliant')
        group_by = options.get('group_by')
        selected_fields = options.get('selected_fields') or []

        query = db.session.query(ReviewResult).options(
            joinedload(ReviewResult.normalized_rule),
            joinedload(ReviewResult.compliance_rule),
            joinedload(ReviewResult.profile)
        ).filter(ReviewResult.review_session_id == review_session_id)
        if source_file:
            query = query.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
        if include_compliant_opt is False:
            query = query.filter(ReviewResult.status == 'non_compliant')

        if group_by == 'severity':
            query = query.order_by(ReviewResult.severity.asc(), ReviewResult.id.asc())
        elif group_by == 'rule':
            query = query.join(ComplianceRule).order_by(ComplianceRule.rule_name.asc(), ReviewResult.id.asc())
        else:
            query = query.join(NormalizedRule).order_by(NormalizedRule.source_file.asc(), ReviewResult.id.asc())

        has_any = db.session.query(query.exists()).scalar()
        if not has_any:
            raise ValueError("No results found for the specified review session")

        def build_base_row(r: ReviewResult) -> Dict[str, Any]:
            nr = r.normalized_rule
            safe = lambda name: getattr(nr, name, '')
            return {
                'Rule_ID': safe('id'),
                'Source_File': safe('source_file'),
                'Rule_Name': safe('rule_name'),
                'Action': safe('action'),
                'Protocol': safe('protocol'),
                'Source_IP': safe('source_ip'),
                'Source_Zone': safe('source_zone'),
                'Source_Port': safe('source_port'),
                'Dest_IP': safe('dest_ip'),
                'Dest_Zone': safe('dest_zone'),
                'Dest_Port': safe('dest_port'),
                'Service_Name': safe('service_name'),
                'Compliance_Status': (r.status or '').replace('_', ' ').title(),
                'Severity': r.severity,
                'Compliance_Rule_Name': r.compliance_rule.rule_name if r.compliance_rule else ''
            }

        if selected_fields:
            headers = selected_fields
        else:
            first_row = query.limit(1).first()
            if not first_row:
                raise ValueError("No results found for the specified review session")
            headers = list(build_base_row(first_row).keys())

        wb = Workbook(write_only=True)
        ws_by_group: Dict[str, Any] = {}
        used_titles: Dict[str, int] = {}

        def safe_sheet_title(name: str) -> str:
            base = (name or 'Unknown').replace('/', '_').replace('\\', '_')
            base = ''.join(ch for ch in base if ch not in '[]:*?/\\')
            base = base.strip() or 'Unknown'
            base = base[:31]
            if base not in used_titles:
                used_titles[base] = 1
                return base
            used_titles[base] += 1
            suffix = f"_{used_titles[base]}"
            return (base[:31 - len(suffix)] + suffix)[:31]

        def group_key(r: ReviewResult) -> str:
            if group_by == 'severity':
                return r.severity or 'Unknown'
            if group_by == 'rule':
                return r.compliance_rule.rule_name if r.compliance_rule else 'Unknown'
            return getattr(r.normalized_rule, 'source_file', None) or 'Unknown'

        def ensure_ws(key: str):
            ws = ws_by_group.get(key)
            if ws:
                return ws
            ws = wb.create_sheet(title=safe_sheet_title(key))
            ws.append(headers)
            ws_by_group[key] = ws
            return ws

        for r in query.yield_per(1000):
            base = build_base_row(r)
            if selected_fields:
                out = {}
                for k in selected_fields:
                    if k in base:
                        out[k] = base[k]
                    else:
                        v = getattr(r.normalized_rule, k, None)
                        out[k] = v if v is not None else ''
            else:
                out = base
            ws = ensure_ws(group_key(r))
            ws.append([out.get(h, '') for h in headers])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        raise Exception(f"Error generating custom Excel export: {str(e)}")

def generate_csv_export(review_session_id: str, source_file: Optional[str] = None, include_compliant: Optional[bool] = None) -> bytes:
    """
    Generate CSV export with compliance status columns.
    If source_file is specified, export only that file's data.
    Otherwise, export all data in a single CSV.
    """
    try:
        # Get review results
        # Optimize: Eager load raw_rule
        query = db.session.query(ReviewResult).options(
            joinedload(ReviewResult.normalized_rule).joinedload(NormalizedRule.raw_rule),
            joinedload(ReviewResult.compliance_rule),
            joinedload(ReviewResult.profile)
        ).filter(ReviewResult.review_session_id == review_session_id)
        
        if source_file:
            query = query.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
        
        results = query.all()
        if include_compliant is not None and include_compliant is False:
            results = [r for r in results if r.status == 'non_compliant']
        
        if not results:
            columns = [
                'Rule_ID','Source_File','Action','Protocol','Source_IP','Source_Zone','Source_Port','Dest_IP','Dest_Zone','Dest_Port',
                'Service_Name','Source_VLAN','Dest_VLAN','Interface','Direction','Logging','Description',
                'Original_Rule','Raw_Text','Raw_Rule_Text','Compliance_Status','Failed_Checks','Severity','Compliance_Rule_Name',
                'Check_Description','Notes','Checked_At','Raw_Rule_Details','Source_PCI_DSS_Categories','Dest_PCI_DSS_Categories'
            ]
            df = pd.DataFrame(columns=columns)
            output = io.StringIO()
            df.to_csv(output, index=False)
            return output.getvalue().encode('utf-8')
        
        # Optimization: Pre-fetch CMDB PCI data for O(1) lookup
        assets = db.session.query(CMDBAsset.ip_address, CMDBAsset.hostname, CMDBAsset.additional_data).all()
        ip_pci_map = {}
        host_pci_map = {}
        
        for ip, hostname, data in assets:
            if not data:
                continue
            try:
                d = json.loads(data)
                cat = d.get('pcidss_asset_category')
                if cat:
                    if ip:
                        ip_pci_map[ip] = cat
                    if hostname:
                        host_pci_map[hostname.lower()] = cat
            except:
                pass

        def get_pci_categories(ip_field: str, hostname_field: str = None) -> str:
            cats = set()
            if ip_field:
                tokens = [t.strip() for t in ip_field.replace(';', ',').split(',')]
                for t in tokens:
                    t_clean = t.lower().replace('host ', '').strip()
                    if t_clean in ip_pci_map:
                        cats.add(ip_pci_map[t_clean])
            if hostname_field:
                h_clean = hostname_field.lower().strip()
                if h_clean in host_pci_map:
                    cats.add(host_pci_map[h_clean])
            return ', '.join(sorted(list(cats)))

        # Create CSV data
        csv_data = []
        for result in results:
            nr = result.normalized_rule
            safe = lambda name: getattr(nr, name, '')
            
            # Use optimized lookup
            src_cats = get_pci_categories(nr.source_ip, nr.source_hostname)
            dst_cats = get_pci_categories(nr.dest_ip, nr.dest_hostname)
            
            row_data = {
                'Rule_ID': safe('id'),
                'Source_File': safe('source_file'),
                'Action': safe('action'),
                'Protocol': safe('protocol'),
                'Source_IP': safe('source_ip'),
                'Source_Zone': safe('source_zone'),
                'Source_Port': safe('source_port'),
                'Dest_IP': safe('dest_ip'),
                'Dest_Zone': safe('dest_zone'),
                'Dest_Port': safe('dest_port'),
                'Service_Name': safe('service_name'),
                'Source_VLAN': safe('source_vlan'),
                'Dest_VLAN': safe('dest_vlan'),
                'Interface': safe('interface'),
                'Direction': safe('direction'),
                'Logging': safe('logging'),
                'Description': safe('description'),
                'Original_Rule': safe('original_rule'),
                'Raw_Text': safe('raw_text'),
                'Raw_Rule_Text': '',
                'Source_PCI_DSS_Categories': src_cats,
                'Dest_PCI_DSS_Categories': dst_cats,
                # New compliance columns
                'Compliance_Status': result.status.replace('_', ' ').title(),
                'Failed_Checks': len(result.failed_checks) if result.failed_checks else 0,
                'Severity': result.severity,
                'Compliance_Rule_Name': result.compliance_rule.rule_name if result.compliance_rule else '',
                'Check_Description': result.compliance_rule.description if result.compliance_rule else '',
                'Notes': result.notes or '',
                'Checked_At': result.checked_at.strftime('%Y-%m-%d %H:%M:%S') if result.checked_at else ''
            }
            try:
                raw_obj = nr.raw_rule
                raw_dict = None
                if raw_obj:
                    row_data['Raw_Rule_Text'] = raw_obj.rule_text or raw_obj.raw_text or ''
                    raw_dict = {
                        'id': raw_obj.id,
                        'rule_name': raw_obj.rule_name,
                        'source_file': raw_obj.source_file,
                        'line_number': raw_obj.file_line_number,
                        'action': raw_obj.action,
                        'protocol': raw_obj.protocol,
                        'source': raw_obj.source,
                        'destination': raw_obj.destination,
                        'service': raw_obj.destination,
                        'raw_text': raw_obj.raw_text
                    }
                else:
                    raw_dict = {
                        'id': safe('id'),
                        'rule_name': safe('rule_name'),
                        'raw_text': safe('raw_text')
                    }
                row_data['Raw_Rule_Details'] = json.dumps(raw_dict)
            except Exception:
                row_data['Raw_Rule_Details'] = ''
                row_data['Raw_Rule_Text'] = ''
            
            csv_data.append(row_data)
        
        # Convert to CSV
        df = pd.DataFrame(csv_data)
        output = io.StringIO()
        df.to_csv(output, index=False)
        return output.getvalue().encode('utf-8')
        
    except Exception as e:
        raise Exception(f"Error generating CSV export: {str(e)}")

def iter_csv_export(review_session_id: str, source_file: Optional[str] = None, include_compliant: Optional[bool] = None) -> Iterator[bytes]:
    include = True if include_compliant is None else bool(include_compliant)
    query = db.session.query(ReviewResult).options(
        joinedload(ReviewResult.normalized_rule).joinedload(NormalizedRule.raw_rule),
        joinedload(ReviewResult.compliance_rule),
        joinedload(ReviewResult.profile)
    ).filter(ReviewResult.review_session_id == review_session_id)

    if not include:
        query = query.filter(ReviewResult.status == 'non_compliant')
    if source_file:
        query = query.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)

    has_any = db.session.query(query.exists()).scalar()
    columns = [
        'Rule_ID','Source_File','Action','Protocol','Source_IP','Source_Zone','Source_Port','Dest_IP','Dest_Zone','Dest_Port',
        'Service_Name','Source_VLAN','Dest_VLAN','Interface','Direction','Logging','Description',
        'Original_Rule','Raw_Text','Raw_Rule_Text','Compliance_Status','Failed_Checks','Severity','Compliance_Rule_Name',
        'Check_Description','Notes','Checked_At','Raw_Rule_Details','Source_PCI_DSS_Categories','Dest_PCI_DSS_Categories'
    ]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    yield buf.getvalue().encode('utf-8')
    buf.seek(0)
    buf.truncate(0)

    if not has_any:
        return

    assets = db.session.query(CMDBAsset.ip_address, CMDBAsset.hostname, CMDBAsset.additional_data).all()
    ip_pci_map: Dict[str, str] = {}
    host_pci_map: Dict[str, str] = {}
    for ip, hostname, data in assets:
        if not data:
            continue
        try:
            d = json.loads(data)
            cat = d.get('pcidss_asset_category')
            if not cat:
                continue
            if ip:
                ip_pci_map[str(ip).strip().lower()] = cat
            if hostname:
                host_pci_map[str(hostname).strip().lower()] = cat
        except Exception:
            continue

    def get_pci_categories(ip_field: str, hostname_field: Optional[str] = None) -> str:
        cats = set()
        if ip_field:
            tokens = [t.strip() for t in str(ip_field).replace(';', ',').split(',') if t.strip()]
            for t in tokens:
                t_clean = t.lower().replace('host ', '').strip()
                cat = ip_pci_map.get(t_clean)
                if cat:
                    cats.add(cat)
        if hostname_field:
            h_clean = str(hostname_field).lower().strip()
            cat = host_pci_map.get(h_clean)
            if cat:
                cats.add(cat)
        return ', '.join(sorted(cats))

    for result in query.yield_per(1000):
        nr = result.normalized_rule
        safe = lambda name: getattr(nr, name, '')

        src_cats = get_pci_categories(getattr(nr, 'source_ip', ''), getattr(nr, 'source_hostname', None))
        dst_cats = get_pci_categories(getattr(nr, 'dest_ip', ''), getattr(nr, 'dest_hostname', None))

        checks_list: List[Dict[str, Any]] = []
        try:
            if isinstance(result.failed_checks, str):
                parsed = json.loads(result.failed_checks) if result.failed_checks else []
            else:
                parsed = result.failed_checks or []
            checks_list = parsed if isinstance(parsed, list) else []
        except Exception:
            checks_list = []

        raw_rule_text = ''
        raw_rule_details = ''
        try:
            raw_obj = getattr(nr, 'raw_rule', None)
            if raw_obj:
                raw_rule_text = raw_obj.rule_text or raw_obj.raw_text or ''
                raw_dict = {
                    'id': raw_obj.id,
                    'rule_name': raw_obj.rule_name,
                    'source_file': raw_obj.source_file,
                    'line_number': raw_obj.file_line_number,
                    'action': raw_obj.action,
                    'protocol': raw_obj.protocol,
                    'source': raw_obj.source,
                    'destination': raw_obj.destination,
                    'service': raw_obj.destination,
                    'raw_text': raw_obj.raw_text
                }
            else:
                raw_dict = {'id': safe('id'), 'rule_name': safe('rule_name'), 'raw_text': safe('raw_text')}
            raw_rule_details = json.dumps(raw_dict)
        except Exception:
            raw_rule_text = ''
            raw_rule_details = ''

        row = [
            safe('id'),
            safe('source_file'),
            safe('action'),
            safe('protocol'),
            safe('source_ip'),
            safe('source_zone'),
            safe('source_port'),
            safe('dest_ip'),
            safe('dest_zone'),
            safe('dest_port'),
            safe('service_name'),
            safe('source_vlan'),
            safe('dest_vlan'),
            safe('interface'),
            safe('direction'),
            safe('logging'),
            safe('description'),
            safe('original_rule'),
            safe('raw_text'),
            raw_rule_text,
            (result.status or '').replace('_', ' ').title(),
            len(checks_list),
            result.severity,
            result.compliance_rule.rule_name if result.compliance_rule else '',
            result.compliance_rule.description if result.compliance_rule else '',
            result.notes or '',
            result.checked_at.strftime('%Y-%m-%d %H:%M:%S') if result.checked_at else '',
            raw_rule_details,
            src_cats,
            dst_cats
        ]

        writer.writerow(row)
        yield buf.getvalue().encode('utf-8')
        buf.seek(0)
        buf.truncate(0)

def iter_csv_export_custom(review_session_id: str, options: Dict[str, Any]) -> Iterator[bytes]:
    source_file = options.get('source_file')
    include_compliant_opt = options.get('include_compliant')
    include = True if include_compliant_opt is None else bool(include_compliant_opt)
    selected_fields = options.get('selected_fields') or []

    query = db.session.query(ReviewResult).options(
        joinedload(ReviewResult.normalized_rule),
        joinedload(ReviewResult.compliance_rule),
        joinedload(ReviewResult.profile)
    ).filter(ReviewResult.review_session_id == review_session_id)
    if source_file:
        query = query.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
    if not include:
        query = query.filter(ReviewResult.status == 'non_compliant')

    base = [
        'Rule_ID','Source_File','Rule_Name','Action','Protocol','Source_IP','Source_Zone','Source_Port','Dest_IP','Dest_Zone',
        'Dest_Port','Service_Name','Source_PCI_DSS_Categories','Dest_PCI_DSS_Categories','Compliance_Status','Severity','Compliance_Rule_Name'
    ]
    headers = selected_fields if selected_fields else base

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    yield buf.getvalue().encode('utf-8')
    buf.seek(0)
    buf.truncate(0)

    assets = db.session.query(CMDBAsset.ip_address, CMDBAsset.hostname, CMDBAsset.additional_data).all()
    ip_pci_map: Dict[str, str] = {}
    host_pci_map: Dict[str, str] = {}
    for ip, hostname, data in assets:
        if not data:
            continue
        try:
            d = json.loads(data)
            cat = d.get('pcidss_asset_category')
            if not cat:
                continue
            if ip:
                ip_pci_map[str(ip).strip().lower()] = cat
            if hostname:
                host_pci_map[str(hostname).strip().lower()] = cat
        except Exception:
            continue

    def get_pci_categories(ip_field: str, hostname_field: Optional[str] = None) -> str:
        cats = set()
        if ip_field:
            tokens = [t.strip() for t in str(ip_field).replace(';', ',').split(',') if t.strip()]
            for t in tokens:
                t_clean = t.lower().replace('host ', '').strip()
                cat = ip_pci_map.get(t_clean)
                if cat:
                    cats.add(cat)
        if hostname_field:
            h_clean = str(hostname_field).lower().strip()
            cat = host_pci_map.get(h_clean)
            if cat:
                cats.add(cat)
        return ', '.join(sorted(cats))

    for r in query.yield_per(1000):
        nr = r.normalized_rule
        base_row = {
            'Rule_ID': getattr(nr, 'id', ''),
            'Source_File': getattr(nr, 'source_file', ''),
            'Rule_Name': getattr(nr, 'rule_name', ''),
            'Action': getattr(nr, 'action', ''),
            'Protocol': getattr(nr, 'protocol', ''),
            'Source_IP': getattr(nr, 'source_ip', ''),
            'Source_Zone': getattr(nr, 'source_zone', ''),
            'Source_Port': getattr(nr, 'source_port', ''),
            'Dest_IP': getattr(nr, 'dest_ip', ''),
            'Dest_Zone': getattr(nr, 'dest_zone', ''),
            'Dest_Port': getattr(nr, 'dest_port', ''),
            'Service_Name': getattr(nr, 'service_name', ''),
            'Source_PCI_DSS_Categories': get_pci_categories(getattr(nr, 'source_ip', ''), getattr(nr, 'source_hostname', None)),
            'Dest_PCI_DSS_Categories': get_pci_categories(getattr(nr, 'dest_ip', ''), getattr(nr, 'dest_hostname', None)),
            'Compliance_Status': (r.status or '').replace('_', ' ').title(),
            'Severity': r.severity,
            'Compliance_Rule_Name': r.compliance_rule.rule_name if r.compliance_rule else ''
        }
        out_row = []
        for h in headers:
            if h in base_row:
                out_row.append(base_row[h])
            else:
                v = getattr(nr, h, None)
                out_row.append(v if v is not None else '')
        writer.writerow(out_row)
        yield buf.getvalue().encode('utf-8')
        buf.seek(0)
        buf.truncate(0)

def generate_csv_export_custom(review_session_id: str, options: Dict[str, Any]) -> bytes:
    try:
        source_file = options.get('source_file')
        include_compliant_opt = options.get('include_compliant')
        selected_fields = options.get('selected_fields') or []
        query = db.session.query(ReviewResult).options(
            joinedload(ReviewResult.normalized_rule),
            joinedload(ReviewResult.compliance_rule),
            joinedload(ReviewResult.profile)
        ).filter(ReviewResult.review_session_id == review_session_id)
        if source_file:
            query = query.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
        results = query.all()
        if include_compliant_opt is False:
            results = [r for r in results if r.status == 'non_compliant']
        rows = []
        for r in results:
            nr = r.normalized_rule
            # Aggregate PCI DSS categories
            try:
                nr_full = nr.to_dict()
                src_cats = list({m.get('pcidss_asset_category') for m in (nr_full.get('source_cmdb_matches') or []) if m.get('pcidss_asset_category')})
                dst_cats = list({m.get('pcidss_asset_category') for m in (nr_full.get('dest_cmdb_matches') or []) if m.get('pcidss_asset_category')})
            except Exception:
                src_cats = []
                dst_cats = []
            base = {
                'Rule_ID': getattr(nr, 'id', ''),
                'Source_File': getattr(nr, 'source_file', ''),
                'Rule_Name': getattr(nr, 'rule_name', ''),
                'Action': getattr(nr, 'action', ''),
                'Protocol': getattr(nr, 'protocol', ''),
                'Source_IP': getattr(nr, 'source_ip', ''),
                'Source_Zone': getattr(nr, 'source_zone', ''),
                'Source_Port': getattr(nr, 'source_port', ''),
                'Dest_IP': getattr(nr, 'dest_ip', ''),
                'Dest_Zone': getattr(nr, 'dest_zone', ''),
                'Dest_Port': getattr(nr, 'dest_port', ''),
                'Service_Name': getattr(nr, 'service_name', ''),
                'Source_PCI_DSS_Categories': ', '.join(src_cats),
                'Dest_PCI_DSS_Categories': ', '.join(dst_cats),
                'Compliance_Status': r.status.replace('_', ' ').title(),
                'Severity': r.severity,
                'Compliance_Rule_Name': r.compliance_rule.rule_name if r.compliance_rule else ''
            }
            if selected_fields:
                filtered = {}
                for key in selected_fields:
                    if key in base:
                        filtered[key] = base[key]
                    else:
                        v = getattr(nr, key, None)
                        if v is not None:
                            filtered[key] = v
                rows.append(filtered)
            else:
                rows.append(base)
        df = pd.DataFrame(rows)
        output = io.StringIO()
        df.to_csv(output, index=False)
        return output.getvalue().encode('utf-8')
    except Exception as e:
        raise Exception(f"Error generating custom CSV export: {str(e)}")

def generate_pdf_export(review_session_id: str, source_file: Optional[str] = None, include_compliant: Optional[bool] = None, group_by: Optional[str] = None) -> bytes:
    try:
        include = True if include_compliant is None else bool(include_compliant)
        base_query = db.session.query(ReviewResult).options(
            joinedload(ReviewResult.normalized_rule),
            joinedload(ReviewResult.compliance_rule),
            joinedload(ReviewResult.profile)
        ).filter(ReviewResult.review_session_id == review_session_id)
        if source_file:
            base_query = base_query.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
        if not include:
            base_query = base_query.filter(ReviewResult.status == 'non_compliant')

        total = base_query.count()
        if total == 0:
            raise ValueError("No results found for the specified review session")

        nonc_query = base_query.filter(ReviewResult.status == 'non_compliant')
        non_compliant_count = nonc_query.count()
        compliant_count = (total - non_compliant_count) if include else 0
        compliance_percentage = round((compliant_count / total) * 100, 2) if total else 0

        severity_rows = db.session.query(ReviewResult.severity, func.count(ReviewResult.id)).select_from(ReviewResult).filter(
            ReviewResult.review_session_id == review_session_id
        )
        if source_file:
            severity_rows = severity_rows.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
        if not include:
            severity_rows = severity_rows.filter(ReviewResult.status == 'non_compliant')
        severity_rows = severity_rows.group_by(ReviewResult.severity).all()
        severity_counts = {k or 'Unknown': int(v) for k, v in severity_rows}

        top_rule_rows = db.session.query(ComplianceRule.rule_name, func.count(ReviewResult.id)).select_from(ReviewResult).join(
            ComplianceRule, ReviewResult.compliance_rule_id == ComplianceRule.id
        ).filter(
            ReviewResult.review_session_id == review_session_id,
            ReviewResult.status == 'non_compliant'
        )
        if source_file:
            top_rule_rows = top_rule_rows.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
        top_rule_rows = top_rule_rows.group_by(ComplianceRule.rule_name).order_by(func.count(ReviewResult.id).desc()).limit(10).all()

        top_file_rows = []
        if group_by and group_by not in ('severity', 'rule'):
            top_file_rows = db.session.query(NormalizedRule.source_file, func.count(ReviewResult.id)).select_from(ReviewResult).join(
                NormalizedRule, ReviewResult.normalized_rule_id == NormalizedRule.id
            ).filter(
                ReviewResult.review_session_id == review_session_id,
                ReviewResult.status == 'non_compliant'
            )
            if source_file:
                top_file_rows = top_file_rows.filter(NormalizedRule.source_file == source_file)
            top_file_rows = top_file_rows.group_by(NormalizedRule.source_file).order_by(func.count(ReviewResult.id).desc()).limit(10).all()

        meta = get_export_metadata(review_session_id)

        def build_row(r: ReviewResult) -> Dict[str, Any]:
            nr = r.normalized_rule
            return {
                'Rule_ID': getattr(nr, 'id', ''),
                'Source_File': getattr(nr, 'source_file', ''),
                'Rule_Name': getattr(nr, 'rule_name', ''),
                'Action': getattr(nr, 'action', ''),
                'Protocol': getattr(nr, 'protocol', ''),
                'Source_IP': getattr(nr, 'source_ip', ''),
                'Dest_IP': getattr(nr, 'dest_ip', ''),
                'Service_Name': getattr(nr, 'service_name', ''),
                'Compliance_Status': (r.status or '').replace('_', ' ').title(),
                'Severity': r.severity or '',
                'Compliance_Rule_Name': r.compliance_rule.rule_name if r.compliance_rule else ''
            }

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        elems: List[Any] = []

        elems.append(Paragraph('Review Results Export', styles['Title']))
        elems.append(Spacer(1, 8))
        elems.append(Paragraph(f"Session: {meta.get('review_session_id','')}", styles['Normal']))
        elems.append(Paragraph(f"Profile: {meta.get('profile_name','')} ({meta.get('compliance_framework','')})", styles['Normal']))
        elems.append(Paragraph(f"Generated: {meta.get('export_generated_at','')}", styles['Normal']))
        elems.append(Paragraph(f"Total: {total} | Compliant: {compliant_count} | Non-Compliant: {non_compliant_count} | Compliance%: {compliance_percentage}", styles['Normal']))
        if source_file:
            elems.append(Paragraph(f"Source File Filter: {source_file}", styles['Normal']))
        if not include:
            elems.append(Paragraph("Filter: Non-Compliant Only", styles['Normal']))
        elems.append(Spacer(1, 12))

        elems.append(Paragraph('Severity Breakdown', styles['Heading2']))
        sev_table_data = [['Severity', 'Count']]
        for sev in ['Critical', 'High', 'Medium', 'Low', 'Unknown']:
            if sev in severity_counts:
                sev_table_data.append([sev, str(severity_counts[sev])])
        sev_table = Table(sev_table_data, repeatRows=1)
        sev_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elems.append(sev_table)
        elems.append(Spacer(1, 12))

        elems.append(Paragraph('Top Violations', styles['Heading2']))
        top_table_data = [['Rule Name', 'Violations']]
        for name, cnt in top_rule_rows:
            top_table_data.append([str(name or ''), str(int(cnt or 0))])
        top_table = Table(top_table_data, repeatRows=1)
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elems.append(top_table)
        elems.append(PageBreak())

        max_rows_per_section = 100
        sections: List[Dict[str, Any]] = [{'title': 'Results', 'filter': None}]
        if group_by == 'severity':
            vals = [k for k in ['Critical', 'High', 'Medium', 'Low', 'Unknown'] if k in severity_counts]
            sections = [{'title': v, 'filter': ('severity', v)} for v in vals]
        elif group_by == 'rule':
            sections = [{'title': str(name or 'Unknown'), 'filter': ('rule', str(name or 'Unknown'))} for name, _ in top_rule_rows] or sections
        elif group_by:
            sections = [{'title': str(sf or 'Unknown'), 'filter': ('source_file', str(sf or 'Unknown'))} for sf, _ in top_file_rows] or sections

        for si, sec in enumerate(sections):
            if si > 0:
                elems.append(PageBreak())
            elems.append(Paragraph(str(sec['title']), styles['Heading2']))
            elems.append(Spacer(1, 8))

            q = base_query
            f = sec['filter']
            if f:
                if f[0] == 'severity':
                    q = q.filter(ReviewResult.severity == f[1])
                elif f[0] == 'rule':
                    q = q.join(ComplianceRule).filter(ComplianceRule.rule_name == f[1])
                elif f[0] == 'source_file':
                    q = q.join(NormalizedRule).filter(NormalizedRule.source_file == f[1])
            rows = [build_row(r) for r in q.order_by(ReviewResult.id.asc()).limit(max_rows_per_section).all()]
            if not rows:
                elems.append(Paragraph('No results found for this section.', styles['Normal']))
                continue
            cols = list(rows[0].keys())
            table_data = [cols] + [[str(row.get(c, '') if row.get(c, '') is not None else '') for c in cols] for row in rows]
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elems.append(table)
            elems.append(Spacer(1, 8))
            elems.append(Paragraph(f"Showing first {len(rows)} rows (max {max_rows_per_section}) for this section.", styles['Normal']))

        doc.build(elems)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        raise Exception(f"Error generating PDF export: {str(e)}")

def generate_pdf_export_custom(review_session_id: str, options: Dict[str, Any]) -> bytes:
    try:
        source_file = options.get('source_file')
        include_compliant_opt = options.get('include_compliant')
        group_by = options.get('group_by')
        selected_fields = options.get('selected_fields') or []
        include_sections = options.get('include_sections') or []

        include = True if include_compliant_opt is None else bool(include_compliant_opt)
        base_query = db.session.query(ReviewResult).options(
            joinedload(ReviewResult.normalized_rule),
            joinedload(ReviewResult.compliance_rule),
            joinedload(ReviewResult.profile)
        ).filter(ReviewResult.review_session_id == review_session_id)
        if source_file:
            base_query = base_query.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
        if not include:
            base_query = base_query.filter(ReviewResult.status == 'non_compliant')

        total = base_query.count()
        if total == 0:
            raise ValueError("No results found for the specified review session")

        nonc_query = base_query.filter(ReviewResult.status == 'non_compliant')
        non_compliant_count = nonc_query.count()
        compliant_count = (total - non_compliant_count) if include else 0
        compliance_percentage = round((compliant_count / total) * 100, 2) if total else 0

        meta = get_export_metadata(review_session_id)

        def build_row(r: ReviewResult) -> Dict[str, Any]:
            nr = r.normalized_rule
            base = {
                'Rule_ID': getattr(nr, 'id', ''),
                'Source_File': getattr(nr, 'source_file', ''),
                'Rule_Name': getattr(nr, 'rule_name', ''),
                'Action': getattr(nr, 'action', ''),
                'Protocol': getattr(nr, 'protocol', ''),
                'Source_IP': getattr(nr, 'source_ip', ''),
                'Dest_IP': getattr(nr, 'dest_ip', ''),
                'Service_Name': getattr(nr, 'service_name', ''),
                'Compliance_Status': (r.status or '').replace('_', ' ').title(),
                'Severity': r.severity or '',
                'Compliance_Rule_Name': r.compliance_rule.rule_name if r.compliance_rule else ''
            }
            if selected_fields:
                filtered: Dict[str, Any] = {}
                for key in selected_fields:
                    if key in base:
                        filtered[key] = base[key]
                    else:
                        v = getattr(nr, key, None)
                        if v is not None:
                            filtered[key] = v
                return filtered
            return base

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        elems: List[Any] = []

        if not include_sections or 'summary' in include_sections:
            elems.append(Paragraph('Review Results Export', styles['Title']))
            elems.append(Spacer(1, 8))
            elems.append(Paragraph(f"Session: {meta.get('review_session_id','')}", styles['Normal']))
            elems.append(Paragraph(f"Profile: {meta.get('profile_name','')} ({meta.get('compliance_framework','')})", styles['Normal']))
            elems.append(Paragraph(f"Generated: {meta.get('export_generated_at','')}", styles['Normal']))
            elems.append(Paragraph(f"Total: {total} | Compliant: {compliant_count} | Non-Compliant: {non_compliant_count} | Compliance%: {compliance_percentage}", styles['Normal']))
            if source_file:
                elems.append(Paragraph(f"Source File Filter: {source_file}", styles['Normal']))
            if not include:
                elems.append(Paragraph("Filter: Non-Compliant Only", styles['Normal']))
            elems.append(Spacer(1, 12))

        if not include_sections or 'details' in include_sections:
            severity_rows = db.session.query(ReviewResult.severity, func.count(ReviewResult.id)).select_from(ReviewResult).filter(
                ReviewResult.review_session_id == review_session_id
            )
            if source_file:
                severity_rows = severity_rows.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
            if not include:
                severity_rows = severity_rows.filter(ReviewResult.status == 'non_compliant')
            severity_rows = severity_rows.group_by(ReviewResult.severity).all()
            severity_counts = {k or 'Unknown': int(v) for k, v in severity_rows}

            top_rule_rows = db.session.query(ComplianceRule.rule_name, func.count(ReviewResult.id)).select_from(ReviewResult).join(
                ComplianceRule, ReviewResult.compliance_rule_id == ComplianceRule.id
            ).filter(
                ReviewResult.review_session_id == review_session_id,
                ReviewResult.status == 'non_compliant'
            )
            if source_file:
                top_rule_rows = top_rule_rows.join(NormalizedRule).filter(NormalizedRule.source_file == source_file)
            top_rule_rows = top_rule_rows.group_by(ComplianceRule.rule_name).order_by(func.count(ReviewResult.id).desc()).limit(10).all()

            top_file_rows = []
            if group_by and group_by not in ('severity', 'rule'):
                top_file_rows = db.session.query(NormalizedRule.source_file, func.count(ReviewResult.id)).select_from(ReviewResult).join(
                    NormalizedRule, ReviewResult.normalized_rule_id == NormalizedRule.id
                ).filter(
                    ReviewResult.review_session_id == review_session_id,
                    ReviewResult.status == 'non_compliant'
                )
                if source_file:
                    top_file_rows = top_file_rows.filter(NormalizedRule.source_file == source_file)
                top_file_rows = top_file_rows.group_by(NormalizedRule.source_file).order_by(func.count(ReviewResult.id).desc()).limit(10).all()

            max_rows_per_section = 100
            sections: List[Dict[str, Any]] = [{'title': 'Results', 'filter': None}]
            if group_by == 'severity':
                vals = [k for k in ['Critical', 'High', 'Medium', 'Low', 'Unknown'] if k in severity_counts]
                sections = [{'title': v, 'filter': ('severity', v)} for v in vals]
            elif group_by == 'rule':
                sections = [{'title': str(name or 'Unknown'), 'filter': ('rule', str(name or 'Unknown'))} for name, _ in top_rule_rows] or sections
            elif group_by:
                sections = [{'title': str(sf or 'Unknown'), 'filter': ('source_file', str(sf or 'Unknown'))} for sf, _ in top_file_rows] or sections

            for si, sec in enumerate(sections):
                if si > 0:
                    elems.append(PageBreak())
                if group_by:
                    elems.append(Paragraph(str(sec['title']), styles['Heading2']))
                    elems.append(Spacer(1, 8))

                q = base_query
                f = sec['filter']
                if f:
                    if f[0] == 'severity':
                        q = q.filter(ReviewResult.severity == f[1])
                    elif f[0] == 'rule':
                        q = q.join(ComplianceRule).filter(ComplianceRule.rule_name == f[1])
                    elif f[0] == 'source_file':
                        q = q.join(NormalizedRule).filter(NormalizedRule.source_file == f[1])

                rows = [build_row(r) for r in q.order_by(ReviewResult.id.asc()).limit(max_rows_per_section).all()]
                if not rows:
                    elems.append(Paragraph('No results found for this section.', styles['Normal']))
                    continue
                cols = list(rows[0].keys())
                table_data = [cols] + [[str(row.get(c, '') if row.get(c, '') is not None else '') for c in cols] for row in rows]
                table = Table(table_data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elems.append(table)
                elems.append(Spacer(1, 8))
                elems.append(Paragraph(f"Showing first {len(rows)} rows (max {max_rows_per_section}) for this section.", styles['Normal']))

        doc.build(elems)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        raise Exception(f"Error generating custom PDF export: {str(e)}")

def get_available_source_files(review_session_id: str) -> List[str]:
    """
    Get list of source files available for export in a review session.
    """
    try:
        files = db.session.query(NormalizedRule.source_file).distinct().join(
            ReviewResult, ReviewResult.normalized_rule_id == NormalizedRule.id
        ).filter(ReviewResult.review_session_id == review_session_id).all()
        
        return [file[0] for file in files]
        
    except Exception as e:
        raise Exception(f"Error getting source files: {str(e)}")

def get_export_metadata(review_session_id: str) -> Dict[str, Any]:
    """
    Get metadata about the export (file counts, review info, etc.)
    """
    try:
        # Get review session info
        first_result = db.session.query(ReviewResult).options(
            joinedload(ReviewResult.profile)
        ).filter(ReviewResult.review_session_id == review_session_id).first()
        
        if not first_result:
            raise ValueError("Review session not found")
        
        # Get statistics
        total_results = db.session.query(ReviewResult).filter(
            ReviewResult.review_session_id == review_session_id
        ).count()
        
        compliant_count = db.session.query(ReviewResult).filter(
            ReviewResult.review_session_id == review_session_id,
            ReviewResult.status == 'compliant'
        ).count()
        
        source_files = get_available_source_files(review_session_id)
        
        return {
            'review_session_id': review_session_id,
            'profile_name': first_result.profile.profile_name,
            'compliance_framework': first_result.profile.compliance_framework,
            'total_rules': total_results,
            'compliant_rules': compliant_count,
            'non_compliant_rules': total_results - compliant_count,
            'compliance_percentage': round((compliant_count / total_results) * 100, 2) if total_results > 0 else 0,
            'source_files': source_files,
            'export_generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
    except Exception as e:
        raise Exception(f"Error getting export metadata: {str(e)}")
